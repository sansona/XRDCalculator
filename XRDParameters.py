import math, textwrap, openpyxl, sys
import matplotlib.pyplot as plt
from openpyxl import Workbook


def scherrer():
    #Calculates average Scherrer Width given number of peaks and peak position + other necessary information.
    print ('How many calculations will you be doing?')
    isDone = False
    while not isDone:
        try:
            calculations = int(input())
            isDone = True
        except ValueError:
            print('Please enter an integer')

    values = []
    sizes = []
    for i in range(calculations):
        if i == 0:
            print ('What is the shape factor? (dimensionless)')
            k = float(input())
            print ('What is the x-ray wavelength used? (in nm)')
            lamb = float(input())
        print ('\nWhat is the 2Theta value of peak #%d?' % (i+1))
        theta = float(input())
        print ('What is the FWHM of peak #%d? (in degrees 2Theta)' % (i+1))
        beta = float(input())
        print ('What is the instrument broadening factor?')
        broad = float(input())
        print('\nYour crystallite size is ' +  str(k * lamb / ((math.radians(beta - broad))*
                                                             math.cos(math.radians(theta/2)))) + ' nm')
        values.append(k * lamb / ((math.radians(beta - broad)*math.cos(math.radians(theta/2)))))
        if i == calculations - 1:
            print ('Your crystallite sizes are ' + str(values))
            print ('\nYour average crystallite size is ' + str(sum(values)/calculations) + ' nm')


def dSpacing():
    #Calculates and displays list of d-spacings given user inputs
    print ('How many calculations will you be doing?')
    isDone = False
    while not isDone:
        try:
            calculations = int(input())
            isDone = True
        except ValueError:
            print('Please enter an integer')

    values = []
    for i in range(calculations):
        if i == 0:
            print ('What is the x-ray wavelength used? (in nm)')
            lamb = float(input())
        print('\nWhat is the 2Theta value of peak #%d?' % (i+1))
        theta = float(input())
        print('\nYour d-spacing is ' + str(lamb / (2*math.sin(math.radians(theta/2)))) + ' nm')
        values.append(lamb / (2*math.sin(math.radians(theta/2))))
    print ('\nYour d-spacings are: ' + str(values))


def graph():
    #Plots 2Theta graph given .xlsx file with 2Theta in B and Intensity in C.
    print ('This function requires the spreadsheets to be in .xlsx format and in the same folder as the program. \n')
    print ('How many plots would you like plotted?')
    
    isDone = False
    while not isDone:
        try:
            numberPlots = int(input())
            isDone = True
        except ValueError:
            print('Please enter an integer')
            
    for i in range(1, int(numberPlots)+1):
        #Sets up title, range of values plotted, and filename
        while i ==1:
            title = str(input('What would you like to title the graph? (Press enter if no title is desired) \n'))
            fileName = str(input('What would you like to name the file? \n'))
            
            print ('Please enter the beginning column number: ')
            isDone = False
            while not isDone:
                try:
                    beginning = int(input())
                    isDone = True
                except ValueError:
                    print('Please enter an integer')

            print ('Please enter the ending column number: ')
            isDone = False
            while not isDone:
                try:
                    ending = int(input())
                    isDone = True
                except ValueError:
                    print('Please enter an integer')
            break

        #Sets variable of spreadsheet opening
        spreadsheet = str(input('What is the name of file #%d? (without the .xlsx) \n' % (i)))
        xrd = openpyxl.load_workbook(spreadsheet + '.xlsx')
        sheet = xrd.get_sheet_by_name(spreadsheet)
        theta = []
        intense = []

        for i in range(beginning, ending):
            twoTheta = sheet.cell(row = i, column = 2).value
            theta.append(twoTheta)
            intensity = sheet.cell(row = i, column = 3).value
            intense.append(intensity)

        normIntense = [float(i)/max(intense) for i in intense]

        #Plots & formats graph
        fig = plt.figure(i, figsize = (16, 4))
        axes = fig.add_subplot(111)
        for axis in ['top', 'bottom', 'left', 'right']:
            axes.spines[axis].set_linewidth(2)
                
        plt.plot(theta, normIntense, linewidth = 2, label = spreadsheet)
        plt.legend(bbox_to_anchor = (1.05, 1), loc = 10, borderaxespad=0.)
        plt.title(title, fontsize = 24)
        plt.xlabel(r'$2\theta$ (deg)', fontsize = 18, fontweight = 'bold')
        plt.ylabel('Intensity (norm)', fontsize = 18, fontweight = 'bold')

    print('The figure has been saved.')
    plt.savefig(fileName + '.png', bbox_inches = 'tight')
    
    

while True:
    print ('\nWhat would you like to do?\n')
    print ('Press 1 to calculate Scherrer Width')
    print ('Press 2 to calculate d-spacing')
    print ('Press 3 to generate plot')
    print ('Press 0 to exit')
    response = int(input())


    if response == 1:
        scherrer()
    elif response == 2:
        dSpacing()
    elif response == 3:
        graph()
    elif response == 0:
        sys.exit()
    response = str(input('\nPress a to return to the menu, or press anything else to exit \n'))
    if response == 'a':
        True
    else:
        break 
